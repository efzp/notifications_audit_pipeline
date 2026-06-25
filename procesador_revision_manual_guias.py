import base64
import binascii
import io
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.utils.normalization import clean_text, normalize_document, normalize_radicado


PAYLOAD_PATH = Path("payload_revision_manual_guias.json")
EXPECTED_FILE_TYPE = "REVISION_MANUAL_GUIAS"
SHEET_NAME = "Revision manual guias"
HEADER_ROW = 2

REQUIRED_FIELDS = {
    "tipo_archivo",
    "nombre_archivo",
    "ruta_sharepoint",
    "identifier",
    "file_content_base64",
}

EXPECTED_HEADERS = [
    "id_notificacion_esperada",
    "nombre_archivo_notificacion_esperada",
    "numero_radicado_normalizado",
    "cedula_normalizada",
    "sala",
    "fecha_audiencia",
    "tipo_destinatario",
    "nombre_entidad",
    "correo_o_guia_entidad",
    "correo_normalizado_entidad",
    "fecha_revision",
    "correo_o_guia_reportado",
    "fecha_envio_reportada",
    "fecha_recibido_reportada",
    "pestana_nombre",
    "comentarios_excel",
    "cumplimiento",
    "cumplimiento_extemporaneo",
    "observaciones",
    "revisado_por",
]

HUMAN_EDIT_FIELDS = {
    "cumplimiento",
    "cumplimiento_extemporaneo",
    "observaciones",
    "revisado_por",
}


def load_payload(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo: {path}")

    with path.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    return validate_payload(payload)


def validate_payload(payload: dict[str, Any]) -> dict[str, Any]:
    missing = sorted(field for field in REQUIRED_FIELDS if not payload.get(field))
    if missing:
        raise ValueError(f"Faltan campos obligatorios en el payload: {', '.join(missing)}")

    if payload["tipo_archivo"] != EXPECTED_FILE_TYPE:
        raise ValueError(
            f"tipo_archivo debe ser {EXPECTED_FILE_TYPE}, recibido: {payload['tipo_archivo']}"
        )

    return payload


def decode_file(payload: dict[str, Any]) -> bytes:
    raw_content = payload["file_content_base64"]
    if isinstance(raw_content, str) and "," in raw_content[:100]:
        raw_content = raw_content.split(",", 1)[1]

    try:
        return base64.b64decode(raw_content, validate=False)
    except (binascii.Error, ValueError, TypeError) as exc:
        raise ValueError("file_content_base64 no es un Base64 valido") from exc


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "_", str(value or "").strip().lower())


def parse_bool_cell(value: Any, field_name: str, row_number: int) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)) and value in (0, 1):
        return int(value)
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "t", "si", "s", "yes", "y"}:
        return 1
    if normalized in {"0", "false", "f", "no", "n"}:
        return 0

    raise ValueError(
        f"{field_name} debe ser 0/1, SI/NO o TRUE/FALSE en fila {row_number}"
    )


def parse_int_cell(value: Any, field_name: str, row_number: int) -> int:
    if value in (None, ""):
        raise ValueError(f"{field_name} es obligatorio en fila {row_number}")
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} debe ser numerico en fila {row_number}") from exc


def row_has_human_edit(row: dict[str, Any]) -> bool:
    return any(row.get(field_name) not in (None, "") for field_name in HUMAN_EDIT_FIELDS)


def extract_revision_rows(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"El archivo debe contener la hoja: {SHEET_NAME}")

    worksheet = workbook[SHEET_NAME]
    raw_headers = [
        normalize_header(cell.value)
        for cell in next(
            worksheet.iter_rows(
                min_row=HEADER_ROW,
                max_row=HEADER_ROW,
            )
        )
    ]
    missing_headers = [
        header for header in EXPECTED_HEADERS if header not in raw_headers
    ]
    if missing_headers:
        raise ValueError(
            "Faltan columnas esperadas en la hoja de revision manual: "
            + ", ".join(missing_headers)
        )

    rows = []
    for row_number, cells in enumerate(
        worksheet.iter_rows(min_row=HEADER_ROW + 1),
        start=HEADER_ROW + 1,
    ):
        values = {
            header: cell.value
            for header, cell in zip(raw_headers, cells)
            if header
        }
        if not row_has_human_edit(values):
            continue

        cumplimiento = parse_bool_cell(
            values.get("cumplimiento"),
            "cumplimiento",
            row_number,
        )
        cumplimiento_extemporaneo = parse_bool_cell(
            values.get("cumplimiento_extemporaneo"),
            "cumplimiento_extemporaneo",
            row_number,
        )
        if cumplimiento == 1 and cumplimiento_extemporaneo == 1:
            raise ValueError(
                "cumplimiento y cumplimiento_extemporaneo no pueden ser ambos 1 "
                f"en fila {row_number}"
            )

        rows.append(
            {
                "numero_linea_excel": row_number,
                "id_notificacion_esperada": parse_int_cell(
                    values.get("id_notificacion_esperada"),
                    "id_notificacion_esperada",
                    row_number,
                ),
                "numero_radicado_normalizado": normalize_radicado(
                    values.get("numero_radicado_normalizado")
                ),
                "cedula_normalizada": normalize_document(
                    values.get("cedula_normalizada")
                ),
                "tipo_destinatario": (
                    clean_text(values.get("tipo_destinatario")) or ""
                ).upper(),
                "cumplimiento": cumplimiento,
                "cumplimiento_extemporaneo": cumplimiento_extemporaneo,
                "observaciones": clean_text(values.get("observaciones")),
                "revisado_por": clean_text(values.get("revisado_por")),
            }
        )

    for row in rows:
        if not row["numero_radicado_normalizado"]:
            raise ValueError(
                "numero_radicado_normalizado es obligatorio "
                f"en fila {row['numero_linea_excel']}"
            )
        if not row["tipo_destinatario"]:
            raise ValueError(
                "tipo_destinatario es obligatorio "
                f"en fila {row['numero_linea_excel']}"
            )

    return rows


def process_payload_data(payload: dict[str, Any]) -> dict[str, Any]:
    payload = validate_payload(payload)
    content = decode_file(payload)
    revision_rows = extract_revision_rows(content)

    return {
        "status": "OK",
        "tipo_archivo": payload["tipo_archivo"],
        "nombre_archivo": payload["nombre_archivo"],
        "ruta_sharepoint": payload["ruta_sharepoint"],
        "tamano_bytes": len(content),
        "total_filas_revision_manual_guias": len(revision_rows),
        "tabla_revision_manual_guias": revision_rows,
    }


def process_payload(payload_path: Path = PAYLOAD_PATH) -> dict[str, Any]:
    return process_payload_data(load_payload(payload_path))


def main() -> int:
    payload_path = Path(sys.argv[1]) if len(sys.argv) > 1 else PAYLOAD_PATH

    try:
        result = process_payload(payload_path)
    except Exception as exc:
        print(json.dumps({"status": "ERROR", "mensaje": str(exc)}, ensure_ascii=False, indent=2))
        return 1

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
