import base64
import binascii
import json
import re
import sys
import unicodedata
import zipfile
from datetime import date, datetime
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.utils.datetime import from_excel


PAYLOAD_PATH = Path("payload.json")

REQUIRED_FIELDS = {
    "tipo_archivo",
    "nombre_archivo",
    "ruta_sharepoint",
    "identifier",
    "file_content_base64",
}


MONTHS = {
    "ene": 1,
    "enero": 1,
    "feb": 2,
    "febrero": 2,
    "mar": 3,
    "marzo": 3,
    "abr": 4,
    "abril": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "junio": 6,
    "jul": 7,
    "julio": 7,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "sept": 9,
    "septiembre": 9,
    "oct": 10,
    "octubre": 10,
    "nov": 11,
    "noviembre": 11,
    "dic": 12,
    "diciembre": 12,
}

DATE_PATTERNS = [
    re.compile(r"\b(?P<day>\d{1,2})[/-](?P<month>\d{1,2})[/-](?P<year>\d{2,4})\b"),
    re.compile(r"\b(?P<year>\d{4})[/-](?P<month>\d{1,2})[/-](?P<day>\d{1,2})\b"),
    re.compile(
        r"\b(?P<day>\d{1,2})\s*(?:de\s*)?"
        r"(?P<month_name>ene(?:ro)?|feb(?:rero)?|mar(?:zo)?|abr(?:il)?|may(?:o)?|"
        r"jun(?:io)?|jul(?:io)?|ago(?:sto)?|sep(?:t|tiembre)?|oct(?:ubre)?|"
        r"nov(?:iembre)?|dic(?:iembre)?)"
        r"\s*(?:de\s*)?(?P<year>\d{2,4})\b",
        re.IGNORECASE,
    ),
]
EMAIL_PATTERN = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)

DETAIL_HEADER_REQUIRED_TERMS = {
    "NO RADICADO",
    "CEDULA",
    "NOMBRE DEL PACIENTE",
    "FECHA DE ENVIO",
    "FECHA DE RECIBIDO",
}

REQUIRED_STRUCTURE = {
    "metadata_hoja": [
        "hoja_trabajo_sala",
        "hoja_trabajo_fecha_audiencia",
    ],
    "encabezado_wide": [
        "regional",
        "pacientes",
        "empleador",
        "remitente",
        "eps",
        "afp",
        "arl",
        "aseguradoras",
    ],
    "encabezado_detalle": [
        "numero_radicado",
        "cedula",
        "nombre_del_paciente",
        "fecha_de_envio",
        "fecha_de_recibido",
    ],
}

STRUCTURE_THRESHOLD = 0.90
FUZZY_MATCH_THRESHOLD = 0.86
NOTIFICATION_ENTITIES = [
    "regional",
    "pacientes",
    "empleador",
    "remitente",
    "eps",
    "afp",
    "arl",
    "aseguradoras",
]
EXCLUDED_NOTIFICATION_EMAIL_VALUES = {
    "remitente",
    "particular",
    "n_a",
    "na",
    "no_aplica",
    "no_informa",
    "no_informan",
    "no_informado",
    "no_reporta",
    "no_reportan",
    "independiente",
    "no_cuenta",
    "no_refiere",
    "sin_correo",
    "sin_email",
}
EXCLUDED_NOTIFICATION_EMAIL_FUZZY_VALUES = {
    "independiente",
    "no_cuenta",
    "no_refiere",
}
EXCLUDED_NOTIFICATION_EMAIL_FUZZY_PHRASES = {
    "independiente",
    "no_cuenta",
    "no_refiere",
    "no_informa",
}
EXCLUDED_NOTIFICATION_EMAIL_FUZZY_THRESHOLD = 0.88

ROW_METADATA_FIELDS = {
    "pestana_nombre": "pestana_nombre",
    "pestana_sala_normalizada": "pestana_sala_normalizada",
    "hoja_trabajo_sala_normalizada": "hoja_trabajo_sala_normalizada",
}

NOTIFICATION_METADATA_FIELDS = {
    **ROW_METADATA_FIELDS,
    "pestana_fecha": "pestana_fecha",
    "hoja_trabajo_fecha_audiencia": "hoja_trabajo_fecha_audiencia",
}

COLUMN_NAME_ALIASES = {
    "n": "numero",
    "no": "numero",
    "no_radicado": "numero_radicado",
    "numero_radicado": "numero_radicado",
    "cedula": "cedula",
    "nombre_del_paciente": "nombre_del_paciente",
    "fecha": "fecha_pago_dictamen",
    "correo_guia": "correo_guia",
    "correo_regional": "regional_correo",
    "correo_remitente": "remitente_correo",
    "correo_eps": "eps_correo",
    "correo_afp": "afp_correo",
    "correo_arl": "arl_correo",
    "correo_cia": "aseguradoras_correo",
    "cia": "aseguradoras",
}

WIDE_HEADER_ALIASES = {
    "REGIONAL": "regional",
    "PACIENTES": "pacientes",
    "EMPLEADOR": "empleador",
    "REMITENTE": "remitente",
    "EPS": "eps",
    "AFP": "afp",
    "ARL": "arl",
    "ASEGURADORAS": "aseguradoras",
    "ASEGURDADORAS": "aseguradoras",
}


def load_payload(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo: {path}")

    with path.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    return validate_payload(payload)


def validate_payload(payload: dict) -> dict:
    missing = sorted(field for field in REQUIRED_FIELDS if not payload.get(field))
    if missing:
        raise ValueError(f"Faltan campos obligatorios en el payload: {', '.join(missing)}")

    return payload


def decode_file(payload: dict) -> bytes:
    try:
        return base64.b64decode(payload["file_content_base64"], validate=True)
    except (binascii.Error, ValueError) as exc:
        raise ValueError("file_content_base64 no es un Base64 valido") from exc


def validate_xlsx(content: bytes) -> list[str]:
    if not content.startswith(b"PK"):
        raise ValueError("El contenido decodificado no parece ser un archivo XLSX/ZIP")

    try:
        with zipfile.ZipFile(BytesIO(content)) as workbook:
            workbook.testzip()
            names = workbook.namelist()
    except zipfile.BadZipFile as exc:
        raise ValueError("El contenido decodificado no es un XLSX valido") from exc

    if "xl/workbook.xml" not in names:
        raise ValueError("El XLSX no contiene xl/workbook.xml")

    return names


def normalize_year(year: str) -> int:
    value = int(year)
    if value < 100:
        return 2000 + value
    return value


def normalize_sheet_name(name: str) -> str:
    return re.sub(r"\s+", " ", name).strip()


def normalize_db_string(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    clean_value = re.sub(r"[^A-Za-z0-9]+", "_", ascii_value)
    return clean_value.strip("_").lower()


def normalize_label(value: object) -> str:
    if value is None:
        return ""

    normalized = unicodedata.normalize("NFKD", str(value))
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]+", " ", ascii_value.upper()).strip()


def similarity(left: str, right: str) -> float:
    return SequenceMatcher(None, left, right).ratio()


def fuzzy_matches(candidate: str, expected: str, threshold: float = FUZZY_MATCH_THRESHOLD) -> bool:
    if not candidate or not expected:
        return False

    if re.search(rf"\b{re.escape(expected)}\b", candidate):
        return True

    if len(expected) <= 3:
        return False

    return similarity(normalize_db_string(candidate), normalize_db_string(expected)) >= threshold


def best_fuzzy_alias(value: str, aliases: dict[str, str], threshold: float = FUZZY_MATCH_THRESHOLD) -> str | None:
    normalized_value = normalize_label(value)
    value_key = normalize_db_string(normalized_value)
    best_match = None
    best_score = 0.0

    for alias, canonical in aliases.items():
        normalized_alias = normalize_label(alias)
        alias_key = normalize_db_string(normalized_alias)

        if value_key == alias_key:
            return canonical

        if len(normalized_alias) <= 3:
            continue

        score = similarity(value_key, alias_key)
        if score > best_score:
            best_score = score
            best_match = canonical

    if best_score >= threshold:
        return best_match

    return None


def normalize_column_name(value: object) -> str:
    normalized = normalize_db_string(str(value))
    if normalized in COLUMN_NAME_ALIASES:
        return COLUMN_NAME_ALIASES[normalized]

    fuzzy_alias = best_fuzzy_alias(str(value), COLUMN_NAME_ALIASES)
    return fuzzy_alias or normalized


def normalize_wide_header(value: object) -> str | None:
    return best_fuzzy_alias(str(value), WIDE_HEADER_ALIASES)


def label_matches_required_term(label: str, required_term: str) -> bool:
    return fuzzy_matches(label, normalize_label(required_term))


def normalize_room_display(value: object) -> str | None:
    if value is None:
        return None

    room = normalize_sheet_name(str(value))
    if not room:
        return None

    if normalize_label(room).startswith("SALA"):
        return room

    return f"Sala {room}"


def remove_room_prefix(value: str) -> str:
    return normalize_sheet_name(
        re.sub(r"^\s*SALA(?:\s+N(?:O|º|°)?\.?:?)?\s*", "", value, flags=re.IGNORECASE)
    )


def normalize_date_value(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except (TypeError, ValueError, OverflowError):
            return None

    date_result = extract_date_from_text(str(value))
    if date_result:
        return date_result[0]

    return None


def extract_date_from_text(text: str) -> tuple[str, tuple[int, int], str] | None:
    for pattern in DATE_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue

        groups = match.groupdict()
        day = int(groups["day"])
        year = normalize_year(groups["year"])

        if groups.get("month_name"):
            month_key = groups["month_name"].lower().strip(".")
            month = MONTHS[month_key]
        else:
            month = int(groups["month"])

        if not 1 <= day <= 31 or not 1 <= month <= 12:
            continue

        return f"{year:04d}-{month:02d}-{day:02d}", match.span(), match.group(0)

    return None


def extract_room_name(sheet_name: str, date_span: tuple[int, int]) -> str:
    without_date = f"{sheet_name[:date_span[0]]} {sheet_name[date_span[1]:]}"
    without_extra_symbols = re.sub(r"[-_/.,()]+", " ", without_date)
    return remove_room_prefix(without_extra_symbols)


def find_adjacent_value(worksheet, target_label: str, max_row: int = 2) -> tuple[object, str] | tuple[None, None]:
    target = normalize_label(target_label)

    for row in worksheet.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if normalize_label(cell.value) != target:
                continue

            adjacent_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
            return adjacent_cell.value, adjacent_cell.coordinate

    return None, None


def extract_sheet_header_metadata(worksheet) -> dict:
    raw_room, room_cell = find_adjacent_value(worksheet, "SALA No")
    raw_audience_date, date_cell = find_adjacent_value(worksheet, "FECHA AUDIENCIA")

    room_display = normalize_room_display(raw_room)
    audience_date = normalize_date_value(raw_audience_date)

    return {
        "hoja_trabajo_sala_original": raw_room,
        "hoja_trabajo_sala": room_display,
        "hoja_trabajo_sala_normalizada": normalize_db_string(room_display) if room_display else None,
        "hoja_trabajo_sala_celda": room_cell,
        "hoja_trabajo_fecha_audiencia_original": raw_audience_date.isoformat()
        if isinstance(raw_audience_date, (date, datetime))
        else raw_audience_date,
        "hoja_trabajo_fecha_audiencia": audience_date,
        "hoja_trabajo_fecha_audiencia_celda": date_cell,
    }


def find_wide_header_row(worksheet, max_row: int = 10) -> dict:
    for row in worksheet.iter_rows(min_row=1, max_row=max_row):
        matches = []
        locations = {}

        for cell in row:
            normalized_value = normalize_label(cell.value)
            if not normalized_value:
                continue

            matched_term = normalize_wide_header(normalized_value)
            if not matched_term:
                continue

            matches.append(matched_term)
            locations[matched_term] = cell.coordinate

        if matches:
            return {
                "fila_encabezado_wide": row[0].row,
                "fila_encabezado_wide_terminos": sorted(set(matches)),
                "ubicacion_columnas_wide": dict(sorted(locations.items())),
            }

    return {
        "fila_encabezado_wide": None,
        "fila_encabezado_wide_terminos": [],
        "ubicacion_columnas_wide": {},
    }


def compact_locations(locations: dict[str, list[str]]) -> dict[str, str | list[str]]:
    return {
        key: cells[0] if len(cells) == 1 else cells
        for key, cells in sorted(locations.items())
    }


def get_column_index(coordinate: str) -> int:
    column_letter, _ = coordinate_from_string(coordinate)
    return column_index_from_string(column_letter)


def find_nearest_left_group(column_index: int, wide_locations: dict[str, str]) -> str | None:
    candidates = [
        (group, get_column_index(coordinate))
        for group, coordinate in wide_locations.items()
        if isinstance(coordinate, str) and get_column_index(coordinate) <= column_index
    ]

    if not candidates:
        return None

    return max(candidates, key=lambda item: item[1])[0]


def add_grouped_date_columns(detail_locations: dict, wide_locations: dict[str, str]) -> dict:
    grouped_locations = dict(detail_locations)

    for field in ("fecha_de_envio", "fecha_de_recibido"):
        coordinates = detail_locations.get(field, [])
        if isinstance(coordinates, str):
            coordinates = [coordinates]

        for coordinate in coordinates:
            group = find_nearest_left_group(get_column_index(coordinate), wide_locations)
            if not group:
                continue

            grouped_locations[f"{group}_{field}"] = coordinate

    grouped_locations.pop("fecha_de_envio", None)
    grouped_locations.pop("fecha_de_recibido", None)

    return dict(sorted(grouped_locations.items()))


def add_grouped_email_columns(detail_locations: dict, wide_locations: dict[str, str]) -> dict:
    grouped_locations = dict(detail_locations)
    email_keys = [key for key in detail_locations if "correo" in key]

    for key in email_keys:
        coordinates = detail_locations.get(key, [])
        if isinstance(coordinates, str):
            coordinates = [coordinates]

        grouped_locations.pop(key, None)

        for coordinate in coordinates:
            group = find_nearest_left_group(get_column_index(coordinate), wide_locations)
            if not group:
                continue

            grouped_locations[f"{group}_correo"] = coordinate

    return dict(sorted(grouped_locations.items()))


def find_detail_header_row(worksheet, max_row: int = 10) -> dict:
    for row in worksheet.iter_rows(min_row=1, max_row=max_row):
        locations = {}
        row_labels = []

        for cell in row:
            normalized_value = normalize_label(cell.value)
            if not normalized_value:
                continue

            row_labels.append(normalized_value)
            key = normalize_column_name(normalized_value)
            locations.setdefault(key, []).append(cell.coordinate)

        matched_required_terms = {
            term
            for term in DETAIL_HEADER_REQUIRED_TERMS
            if any(label_matches_required_term(label, term) for label in row_labels)
        }

        if matched_required_terms == DETAIL_HEADER_REQUIRED_TERMS:
            return {
                "fila_encabezado_detalle": row[0].row,
                "fila_encabezado_detalle_terminos": [
                    normalize_column_name(term) for term in sorted(matched_required_terms)
                ],
                "ubicacion_columnas_detalle": compact_locations(locations),
            }

    return {
        "fila_encabezado_detalle": None,
        "fila_encabezado_detalle_terminos": [],
        "ubicacion_columnas_detalle": {},
    }


def validate_sheet_structure(sheet_metadata: dict, threshold: float = STRUCTURE_THRESHOLD) -> dict:
    checks = []
    missing = []

    metadata_locations = {
        "hoja_trabajo_sala": sheet_metadata.get("hoja_trabajo_sala_celda"),
        "hoja_trabajo_fecha_audiencia": sheet_metadata.get("hoja_trabajo_fecha_audiencia_celda"),
    }
    wide_locations = sheet_metadata.get("ubicacion_columnas_wide", {})
    detail_locations = sheet_metadata.get("ubicacion_columnas_detalle", {})

    for group, fields in REQUIRED_STRUCTURE.items():
        for field in fields:
            if group == "metadata_hoja":
                found = bool(sheet_metadata.get(field)) and bool(metadata_locations.get(field))
                location = metadata_locations.get(field)
            elif group == "encabezado_wide":
                found = field in wide_locations
                location = wide_locations.get(field)
            else:
                if field in ("fecha_de_envio", "fecha_de_recibido"):
                    matching_locations = {
                        key: value
                        for key, value in detail_locations.items()
                        if key.endswith(f"_{field}")
                    }
                    found = bool(matching_locations)
                    location = matching_locations or None
                else:
                    found = field in detail_locations
                    location = detail_locations.get(field)

            checks.append(
                {
                    "grupo": group,
                    "campo": field,
                    "encontrado": found,
                    "ubicacion": location,
                }
            )

            if not found:
                missing.append({"grupo": group, "campo": field})

    total = len(checks)
    found_count = sum(1 for check in checks if check["encontrado"])
    compliance = found_count / total if total else 0
    is_valid = compliance >= threshold

    return {
        "estructura_valida": is_valid,
        "estructura_cumplimiento": round(compliance, 4),
        "estructura_umbral": threshold,
        "estructura_campos_esperados": total,
        "estructura_campos_encontrados": found_count,
        "estructura_faltantes": missing,
    }


def build_standard_column_dictionary(sheets_metadata: list[dict]) -> dict:
    standard = {}

    for sheet in sheets_metadata:
        sheet_name = sheet["pestana_nombre"]
        detail_locations = sheet.get("ubicacion_columnas_detalle", {})

        for column_name, location in detail_locations.items():
            entry = standard.setdefault(
                column_name,
                {
                    "nombre_estandar": column_name,
                    "hojas_detectadas": 0,
                    "ubicaciones_detectadas": {},
                    "ubicaciones_por_hoja": {},
                    "requerida_estructura": any(
                        column_name == field or column_name.endswith(f"_{field}")
                        for fields in REQUIRED_STRUCTURE.values()
                        for field in fields
                    ),
                },
            )

            entry["hojas_detectadas"] += 1
            entry["ubicaciones_por_hoja"][sheet_name] = location

            locations = location if isinstance(location, list) else [location]
            for cell in locations:
                entry["ubicaciones_detectadas"][cell] = entry["ubicaciones_detectadas"].get(cell, 0) + 1

    for entry in standard.values():
        entry["ubicaciones_detectadas"] = dict(sorted(entry["ubicaciones_detectadas"].items()))

    return dict(sorted(standard.items()))


def is_notification_column(column_name: str) -> bool:
    return (
        column_name.endswith("_fecha_de_envio")
        or column_name.endswith("_fecha_de_recibido")
        or column_name.endswith("_correo")
    )


def serialize_cell_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, str):
        return normalize_sheet_name(value)

    return value


def serialize_text_value(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return normalize_sheet_name(str(value))


def serialize_date_like_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)):
        normalized_date = normalize_date_value(value)
        return normalized_date if normalized_date else value

    if isinstance(value, str):
        clean_value = normalize_sheet_name(value)
        normalized_date = normalize_date_value(clean_value)
        return normalized_date if normalized_date else clean_value

    return value


def should_skip_notification_email(value: object) -> bool:
    if value is None:
        return False

    raw_value = str(value).strip()
    if not raw_value:
        return True

    normalized_value = normalize_db_string(raw_value)
    if normalized_value in EXCLUDED_NOTIFICATION_EMAIL_VALUES:
        return True

    if "@" not in raw_value and any(
        SequenceMatcher(None, normalized_value, excluded_value).ratio()
        >= EXCLUDED_NOTIFICATION_EMAIL_FUZZY_THRESHOLD
        for excluded_value in EXCLUDED_NOTIFICATION_EMAIL_FUZZY_VALUES
    ):
        return True

    if "@" not in raw_value and should_skip_by_fuzzy_phrase(normalized_value):
        return True

    if "@" not in raw_value and re.fullmatch(r"[\d\s.,/\-]+", raw_value):
        return True

    return False


def should_skip_by_fuzzy_phrase(normalized_value: str) -> bool:
    parts = [part for part in normalized_value.split("_") if part]
    if not parts:
        return False

    for excluded_phrase in EXCLUDED_NOTIFICATION_EMAIL_FUZZY_PHRASES:
        phrase_size = len(excluded_phrase.split("_"))
        if len(parts) < phrase_size:
            continue

        for index in range(len(parts) - phrase_size + 1):
            window = "_".join(parts[index : index + phrase_size])
            if (
                SequenceMatcher(None, window, excluded_phrase).ratio()
                >= EXCLUDED_NOTIFICATION_EMAIL_FUZZY_THRESHOLD
            ):
                return True

    return False


def split_notification_emails(value: object) -> list[object]:
    serialized_value = serialize_cell_value(value)
    if not isinstance(serialized_value, str):
        return [serialized_value]

    matches = EMAIL_PATTERN.findall(serialized_value)
    if not matches:
        return [serialized_value]

    emails = []
    seen = set()
    for email in matches:
        normalized_email = email.strip().lower()
        if normalized_email in seen:
            continue

        emails.append(normalized_email)
        seen.add(normalized_email)

    return emails


def first_coordinate(location: str | list[str]) -> str:
    return location[0] if isinstance(location, list) else location


def pick_metadata(sheet_metadata: dict, fields: dict[str, str]) -> dict:
    return {
        output_name: sheet_metadata[source_name]
        for output_name, source_name in fields.items()
    }


def build_coordinate_indexes(coordinates: list[str]) -> dict[str, int]:
    return {
        coordinate: get_column_index(coordinate)
        for coordinate in coordinates
    }


def iter_sheet_values_by_coordinate(worksheet, start_row: int, coordinates: list[str]):
    coordinate_indexes = build_coordinate_indexes(coordinates)
    unique_column_indexes = sorted(set(coordinate_indexes.values()))
    min_column = min(unique_column_indexes)
    max_column = max(unique_column_indexes)

    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=worksheet.max_row,
        min_col=min_column,
        max_col=max_column,
        values_only=True,
    ):
        values_by_column = {
            column_index: row[column_index - min_column]
            for column_index in unique_column_indexes
        }
        yield {
            coordinate: values_by_column.get(column_index)
            for coordinate, column_index in coordinate_indexes.items()
        }


def build_base_case_columns(detail_locations: dict) -> dict[str, str]:
    return {
        column_name: first_coordinate(location)
        for column_name, location in detail_locations.items()
        if column_name not in {"fecha_de_audiencia", "cedula"}
        and (column_name == "numero_radicado" or not is_notification_column(column_name))
    }


def build_notification_columns(detail_locations: dict) -> dict[str, dict[str, str]]:
    notification_columns = {}

    for entity in NOTIFICATION_ENTITIES:
        entity_columns = {
            "fecha_envio": detail_locations.get(f"{entity}_fecha_de_envio"),
            "fecha_recibido": detail_locations.get(f"{entity}_fecha_de_recibido"),
            "correo": detail_locations.get(f"{entity}_correo"),
        }
        entity_columns = {
            field: first_coordinate(location)
            for field, location in entity_columns.items()
            if location
        }

        if entity_columns:
            notification_columns[entity] = entity_columns

    return notification_columns


def extract_base_case_table(workbook, sheets_metadata: list[dict]) -> list[dict]:
    rows = []

    for sheet_metadata in sheets_metadata:
        if not sheet_metadata["estructura_valida"]:
            continue

        sheet_name = sheet_metadata["pestana_nombre"]
        worksheet = workbook[sheet_name]
        detail_row = sheet_metadata["fila_encabezado_detalle"]
        detail_locations = sheet_metadata["ubicacion_columnas_detalle"]
        base_columns = build_base_case_columns(detail_locations)
        key_coordinate = base_columns.get("numero_radicado")

        if not detail_row or not key_coordinate:
            continue

        coordinates = list(base_columns.values())
        for values_by_coordinate in iter_sheet_values_by_coordinate(worksheet, detail_row + 1, coordinates):
            numero_radicado = values_by_coordinate.get(key_coordinate)
            if numero_radicado in (None, ""):
                continue

            output_row = pick_metadata(sheet_metadata, ROW_METADATA_FIELDS)

            for column_name, coordinate in base_columns.items():
                output_row[column_name] = serialize_cell_value(values_by_coordinate.get(coordinate))

            rows.append(output_row)

    return rows


def extract_notification_table(workbook, sheets_metadata: list[dict]) -> list[dict]:
    rows = []

    for sheet_metadata in sheets_metadata:
        if not sheet_metadata["estructura_valida"]:
            continue

        sheet_name = sheet_metadata["pestana_nombre"]
        worksheet = workbook[sheet_name]
        detail_row = sheet_metadata["fila_encabezado_detalle"]
        detail_locations = sheet_metadata["ubicacion_columnas_detalle"]
        key_coordinate = detail_locations.get("numero_radicado")
        cedula_coordinate = detail_locations.get("cedula")
        notification_columns = build_notification_columns(detail_locations)

        if not detail_row or not key_coordinate or not notification_columns:
            continue

        all_coordinates = [first_coordinate(key_coordinate)]
        if cedula_coordinate:
            all_coordinates.append(first_coordinate(cedula_coordinate))
        for entity_columns in notification_columns.values():
            all_coordinates.extend(entity_columns.values())

        for values_by_coordinate in iter_sheet_values_by_coordinate(
            worksheet,
            detail_row + 1,
            all_coordinates,
        ):
            numero_radicado = values_by_coordinate.get(first_coordinate(key_coordinate))
            if numero_radicado in (None, ""):
                continue

            for entity, entity_columns in notification_columns.items():
                output_row = {
                    **pick_metadata(sheet_metadata, NOTIFICATION_METADATA_FIELDS),
                    "numero_radicado": serialize_text_value(numero_radicado),
                    "cedula": serialize_text_value(
                        values_by_coordinate.get(first_coordinate(cedula_coordinate))
                    )
                    if cedula_coordinate
                    else None,
                    "entidad": entity,
                    "fecha_envio": None,
                    "fecha_recibido": None,
                    "correo": None,
                }

                for field, coordinate in entity_columns.items():
                    value = values_by_coordinate.get(coordinate)
                    if field in {"fecha_envio", "fecha_recibido"}:
                        output_row[field] = serialize_date_like_value(value)
                    elif field == "correo":
                        output_row[field] = serialize_text_value(value)
                    else:
                        output_row[field] = serialize_cell_value(value)

                for correo in split_notification_emails(output_row["correo"]):
                    if should_skip_notification_email(correo):
                        continue

                    notification_row = dict(output_row)
                    notification_row["correo"] = correo
                    rows.append(notification_row)

    return rows


def extract_dated_sheets(workbook) -> list[dict]:
    dated_sheets = []

    for sheet_name in workbook.sheetnames:
        date_result = extract_date_from_text(sheet_name)
        worksheet = workbook[sheet_name]
        header_metadata = extract_sheet_header_metadata(worksheet)
        wide_header_metadata = find_wide_header_row(worksheet)
        detail_header_metadata = find_detail_header_row(worksheet)
        detail_header_metadata["ubicacion_columnas_detalle"] = add_grouped_date_columns(
            detail_header_metadata["ubicacion_columnas_detalle"],
            wide_header_metadata["ubicacion_columnas_wide"],
        )
        detail_header_metadata["ubicacion_columnas_detalle"] = add_grouped_email_columns(
            detail_header_metadata["ubicacion_columnas_detalle"],
            wide_header_metadata["ubicacion_columnas_wide"],
        )

        if not date_result and not header_metadata["hoja_trabajo_fecha_audiencia"]:
            continue

        if date_result:
            extracted_date, date_span, raw_date = date_result
            room_name = extract_room_name(sheet_name, date_span)
        else:
            extracted_date = header_metadata["hoja_trabajo_fecha_audiencia"]
            raw_date = header_metadata["hoja_trabajo_fecha_audiencia_original"]
            room_name = header_metadata["hoja_trabajo_sala"] or normalize_sheet_name(sheet_name)

        sheet_metadata = {
            "pestana_nombre": sheet_name,
            "pestana_nombre_normalizado": normalize_db_string(sheet_name),
            "pestana_fecha": extracted_date,
            "pestana_fecha_original": raw_date,
            "pestana_sala_original": room_name,
            "pestana_sala_normalizada": normalize_db_string(room_name),
            **header_metadata,
            **wide_header_metadata,
            **detail_header_metadata,
        }
        sheet_metadata.update(validate_sheet_structure(sheet_metadata))

        dated_sheets.append(sheet_metadata)

    return dated_sheets


def process_payload_data(payload: dict) -> dict:
    payload = validate_payload(payload)
    content = decode_file(payload)
    entries = validate_xlsx(content)
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    dated_sheets = extract_dated_sheets(workbook)
    standard_column_dictionary = build_standard_column_dictionary(dated_sheets)
    mensaje_error = [
        {
            "pestana": sheet["pestana_nombre"],
            "tipo_error": "estructura_invalida",
            "cumplimiento": sheet["estructura_cumplimiento"],
            "umbral": sheet["estructura_umbral"],
            "campos_esperados": sheet["estructura_campos_esperados"],
            "campos_encontrados": sheet["estructura_campos_encontrados"],
            "faltantes": sheet["estructura_faltantes"],
            "mensaje": "La hoja no cumple con el 90% de la estructura requerida",
        }
        for sheet in dated_sheets
        if not sheet["estructura_valida"]
    ]

    status = "ERROR" if mensaje_error else "OK"
    base_case_table = [] if mensaje_error else extract_base_case_table(workbook, dated_sheets)
    notification_table = [] if mensaje_error else extract_notification_table(workbook, dated_sheets)

    return {
        "status": status,
        "tipo_archivo": payload["tipo_archivo"],
        "nombre_archivo": payload["nombre_archivo"],
        "ruta_sharepoint": payload["ruta_sharepoint"],
        "tamano_bytes": len(content),
        "entradas_xlsx": len(entries),
        "mensaje_error": mensaje_error,
        "diccionario_estandar_columnas": standard_column_dictionary,
        "tabla_casos": base_case_table if not mensaje_error else [],
        "tabla_notificaciones": notification_table,
        "hojas_con_fecha": dated_sheets,
    }


def process_payload(payload_path: Path = PAYLOAD_PATH) -> dict:
    return process_payload_data(load_payload(payload_path))


def main() -> int:
    payload_path = Path(sys.argv[1]) if len(sys.argv) > 1 else PAYLOAD_PATH

    try:
        result = process_payload(payload_path)
    except Exception as exc:
        print(json.dumps({"status": "ERROR", "mensaje": str(exc)}, ensure_ascii=False, indent=2))
        return 1

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
