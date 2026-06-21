import base64
import binascii
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any


PAYLOAD_PATH = Path("payload_calificaciones.json")

REQUIRED_FIELDS = {
    "tipo_archivo",
    "nombre_archivo",
    "ruta_sharepoint",
    "identifier",
    "file_content_base64",
}

EXPECTED_FILE_TYPES = {"SISTEMA_JNC", "CALIFICACIONES_SOFTWARE", "CALIFICACIONES"}
EXPECTED_SHEET_NAME = "Calificaciones"


def load_payload(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo: {path}")

    with path.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    return validate_payload(payload)


def validate_payload(payload: dict[str, Any]) -> dict[str, Any]:
    missing = sorted(field for field in REQUIRED_FIELDS if not payload.get(field))
    if missing:
        raise ValueError(f"Faltan campos obligatorios en el payload: {', '.join(missing)}")

    if payload["tipo_archivo"] not in EXPECTED_FILE_TYPES:
        expected = ", ".join(sorted(EXPECTED_FILE_TYPES))
        raise ValueError(
            f"tipo_archivo debe ser uno de: {expected}. Recibido: {payload['tipo_archivo']}"
        )

    return payload


def decode_file(payload: dict[str, Any]) -> bytes:
    raw_content = payload["file_content_base64"]
    if isinstance(raw_content, str) and "," in raw_content[:100]:
        raw_content = raw_content.split(",", 1)[1]

    try:
        return base64.b64decode(raw_content, validate=False)
    except (binascii.Error, ValueError, TypeError) as exc:
        raise ValueError("file_content_base64 no es un Base64 valido") from exc


def clean_text_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()

    clean_value = re.sub(r"\s+", " ", str(value)).strip()
    if len(clean_value) >= 2 and clean_value[0] == clean_value[-1] == '"':
        clean_value = clean_value[1:-1].strip()
    return clean_value or None


def normalize_column_name(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    clean_value = re.sub(r"[^A-Za-z0-9]+", "_", ascii_value)
    return clean_value.strip("_").lower()


def make_unique_headers(headers: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    unique_headers = []

    for header in headers:
        normalized = normalize_column_name(header) or "columna_sin_nombre"
        seen[normalized] = seen.get(normalized, 0) + 1
        unique_headers.append(
            normalized if seen[normalized] == 1 else f"{normalized}_{seen[normalized]}"
        )

    return unique_headers


def load_calificaciones_xlsx(content: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl no esta instalado. Agregue openpyxl>=3.1.0 a requirements.txt y redeploye la Function."
        ) from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = (
        workbook[EXPECTED_SHEET_NAME]
        if EXPECTED_SHEET_NAME in workbook.sheetnames
        else workbook.worksheets[0]
    )

    row_iter = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = [clean_text_value(value) or "" for value in next(row_iter)]
    except StopIteration as exc:
        raise ValueError("El archivo de calificaciones no contiene encabezados") from exc

    normalized_headers = make_unique_headers(raw_headers)
    rows = []

    for row_number, values in enumerate(row_iter, start=2):
        row = {
            header: clean_text_value(value)
            for header, value in zip(normalized_headers, values)
        }
        if not any(value not in (None, "") for value in row.values()):
            continue

        row["numero_fila_excel"] = row_number
        row["hoja_origen"] = worksheet.title
        rows.append(row)

    return {
        "nombre_hoja": worksheet.title,
        "encabezados_originales": raw_headers,
        "encabezados_normalizados": normalized_headers,
        "total_filas": len(rows),
        "filas": rows,
    }


def process_payload_data(payload: dict[str, Any]) -> dict[str, Any]:
    payload = validate_payload(payload)
    content = decode_file(payload)
    workbook_data = load_calificaciones_xlsx(content)

    return {
        "status": "OK",
        "tipo_archivo": payload["tipo_archivo"],
        "nombre_archivo": payload["nombre_archivo"],
        "ruta_sharepoint": payload["ruta_sharepoint"],
        "tamano_bytes": len(content),
        "nombre_hoja": workbook_data["nombre_hoja"],
        "encabezados_originales": workbook_data["encabezados_originales"],
        "encabezados_normalizados": workbook_data["encabezados_normalizados"],
        "total_filas_calificaciones_software": workbook_data["total_filas"],
        "tabla_calificaciones_software": workbook_data["filas"],
    }


def process_payload(payload_path: Path = PAYLOAD_PATH) -> dict[str, Any]:
    return process_payload_data(load_payload(payload_path))


if __name__ == "__main__":
    try:
        print(json.dumps(process_payload(), ensure_ascii=False, indent=2, default=str))
    except Exception as exc:
        print(json.dumps({"status": "ERROR", "mensaje": str(exc)}, ensure_ascii=False))
        sys.exit(1)
