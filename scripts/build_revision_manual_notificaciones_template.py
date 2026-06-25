from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_PATH = Path("salida") / "template_revision_manual_notificaciones.xlsx"


HEADERS = [
    "id_notificacion_esperada",
    "nombre_archivo_notificacion_esperada",
    "numero_radicado_normalizado",
    "cedula_normalizada",
    "sala",
    "fecha_audiencia",
    "tipo_destinatario",
    "nombre_entidad",
    "correo_o_guia_entidad",
    "correo_normalizado_entidad",
    "fecha_revision",
    "correo_o_guia_reportado",
    "fecha_envio_reportada",
    "fecha_recibido_reportada",
    "pestana_nombre",
    "comentarios_excel",
    "cumplimiento",
    "cumplimiento_extemporaneo",
    "observaciones",
    "revisado_por",
]


EXAMPLE_ROW = [
    12345,
    "Ejemplo archivo de trabajo.xlsx",
    "11001310500120240012300",
    "1020304050",
    "Sala 1",
    "2026-06-18",
    "PACIENTES",
    "Paciente Ejemplo",
    "700123456789",
    None,
    None,
    "700123456789",
    "2026-06-18",
    None,
    "20-06-2026 Sala 1",
    "B12 (EPS) Usuario: comentario de ejemplo",
    1,
    0,
    "Validado manualmente contra evidencia disponible.",
    "usuario@empresa.com",
]


DESCRIPTIONS = {
    "id_notificacion_esperada": "Identificador tecnico preferido. Si viene desde la vista de pendientes, evita ambiguedades.",
    "nombre_archivo_notificacion_esperada": "Nombre del archivo origen de la fila en notificacion_esperada.",
    "numero_radicado_normalizado": "Radicado sin espacios, puntos ni guiones. Obligatorio si no se informa id_notificacion_esperada.",
    "cedula_normalizada": "Cedula solo con digitos. Ayuda a resolver la notificacion esperada.",
    "sala": "Sala tomada de calificacion_sistema_caso.",
    "fecha_audiencia": "Fecha de audiencia tomada de calificacion_sistema_caso o de notificacion_esperada.",
    "tipo_destinatario": "Entidad revisada: PACIENTES, REGIONAL, EMPLEADOR, REMITENTE, EPS, AFP, ARL o ASEGURADORAS.",
    "nombre_entidad": "Nombre de la entidad correspondiente a tipo_destinatario. Evita columnas separadas por EPS, ARL, AFP, etc.",
    "correo_o_guia_entidad": "Correo, guia o dato reportado para la entidad de esta fila.",
    "correo_normalizado_entidad": "Correo normalizado de la entidad si aplica. Para guias puede quedar vacio.",
    "cumplimiento": "Campo humano: 1 si la revision manual valida cumplimiento en plazo; 0 o vacio si no.",
    "cumplimiento_extemporaneo": "Campo humano: 1 si la revision manual valida cumplimiento fuera de plazo; 0 o vacio si no.",
    "fecha_revision": "Se llena automaticamente al cargar/aplicar la revision.",
    "correo_o_guia_reportado": "Valor original en notificacion_esperada.",
    "fecha_envio_reportada": "Fecha de envio reportada en notificacion_esperada.",
    "fecha_recibido_reportada": "Fecha de recibido reportada en notificacion_esperada.",
    "pestana_nombre": "Pestana de origen de la notificacion esperada.",
    "comentarios_excel": "Comentarios extraidos de la fila del caso en el Excel de salas.",
    "observaciones": "Campo libre que llena el humano.",
    "revisado_por": "Usuario, correo o responsable de la revision.",
}


TIPOS_DESTINATARIO = [
    ["tipo_destinatario"],
    ["PACIENTES"],
    ["REGIONAL"],
    ["EMPLEADOR"],
    ["REMITENTE"],
    ["EPS"],
    ["AFP"],
    ["ARL"],
    ["ASEGURADORAS"],
]


BIT_VALUES = [["valor_bit"], [0], [1]]


def style_sheet(ws):
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Template revision manual de notificaciones"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.row_dimensions[1].height = 26

    for cell in ws[2]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cell.value in {
            "cumplimiento",
            "cumplimiento_extemporaneo",
            "observaciones",
            "revisado_por",
        }:
            cell.fill = required_fill
        description = DESCRIPTIONS.get(cell.value)
        if description:
            cell.comment = Comment(description, "Codex")

    widths = {
        "A": 22,
        "B": 40,
        "C": 28,
        "D": 18,
        "E": 18,
        "F": 19,
        "G": 34,
        "H": 28,
        "I": 30,
        "J": 28,
        "K": 20,
        "L": 28,
        "M": 18,
        "N": 19,
        "O": 26,
        "P": 46,
        "Q": 16,
        "R": 24,
        "S": 52,
        "T": 24,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=3, max_row=202, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_validations(wb, ws):
    catalogs = wb["Catalogos"]

    dv_tipo = DataValidation(type="list", formula1="=Catalogos!$E$2:$E$9", allow_blank=True)
    dv_bit = DataValidation(type="list", formula1="=Catalogos!$H$2:$H$3", allow_blank=True)
    dv_date = DataValidation(type="date", operator="between", formula1="DATE(2020,1,1)", formula2="DATE(2035,12,31)", allow_blank=True)

    ws.add_data_validation(dv_tipo)
    ws.add_data_validation(dv_bit)
    ws.add_data_validation(dv_date)

    dv_tipo.add("G3:G202")
    for col in ["Q", "R"]:
        dv_bit.add(f"{col}3:{col}202")
    dv_date.add("F3:F202")
    dv_date.add("K3:K202")
    dv_date.add("M3:N202")

    catalogs.sheet_state = "visible"


def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Revision manual notificaciones"
    catalogs = wb.create_sheet("Catalogos")
    notes = wb.create_sheet("Instrucciones")

    ws.append([""] * len(HEADERS))
    ws.append(HEADERS)
    ws.append(EXAMPLE_ROW)
    for _ in range(199):
        ws.append([None] * len(HEADERS))

    style_sheet(ws)

    last_column = get_column_letter(len(HEADERS))
    table = Table(displayName="RevisionManualNotificaciones", ref=f"A2:{last_column}202")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    catalogs["A1"] = "regla_resultado"
    catalogs["A2"] = "cumplimiento=1 => CUMPLE"
    catalogs["A3"] = "cumplimiento_extemporaneo=1 => FUERA_DE_PLAZO"
    catalogs["A4"] = "ambos vacios/0 => no aplicar"
    catalogs["A5"] = "ambos 1 => error de carga"
    catalogs["E1"] = "tipo_destinatario"
    for index, row in enumerate(TIPOS_DESTINATARIO[1:], start=2):
        catalogs[f"E{index}"] = row[0]
    catalogs["H1"] = "valor_bit"
    catalogs["H2"] = 0
    catalogs["H3"] = 1

    for row in catalogs.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in catalogs[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    catalogs.column_dimensions["A"].width = 28
    catalogs.column_dimensions["B"].width = 24
    catalogs.column_dimensions["C"].width = 60
    catalogs.column_dimensions["E"].width = 24
    catalogs.column_dimensions["H"].width = 12
    catalogs.freeze_panes = "A2"

    notes["A1"] = "Uso recomendado"
    notes["A1"].font = Font(bold=True, size=14, color="1F4E78")
    note_rows = [
        "1. Preferir id_notificacion_esperada cuando venga de la vista de pendientes.",
        "2. Si no hay id, informar radicado, cedula y tipo_destinatario para resolver la notificacion.",
        "3. La consulta base debe traer nombre_entidad, correo_o_guia_entidad y correo_normalizado_entidad segun tipo_destinatario.",
        "4. cumplimiento=1 materializa CUMPLE; cumplimiento_extemporaneo=1 materializa FUERA_DE_PLAZO.",
        "5. No editar directamente resumen_validacion_radicado; debe refrescarse desde los cruces.",
        "6. Las columnas amarillas son las unicas que debe llenar el auditor.",
    ]
    for idx, value in enumerate(note_rows, start=3):
        notes[f"A{idx}"] = value
        notes[f"A{idx}"].alignment = Alignment(wrap_text=True)
    notes.column_dimensions["A"].width = 110
    notes.sheet_view.showGridLines = False

    add_validations(wb, ws)

    ws.conditional_formatting.add(
        "Q3:R202",
        FormulaRule(
            formula=["$Q3=1"],
            fill=PatternFill("solid", fgColor="E2F0D9"),
        ),
    )
    ws.conditional_formatting.add(
        "Q3:R202",
        FormulaRule(
            formula=["$R3=1"],
            fill=PatternFill("solid", fgColor="FCE4D6"),
        ),
    )
    ws.conditional_formatting.add(
        "Q3:R202",
        FormulaRule(
            formula=["AND($Q3=1,$R3=1)"],
            fill=PatternFill("solid", fgColor="F4CCCC"),
        ),
    )

    for sheet in wb.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1

    return wb


def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = create_workbook()
    wb.save(OUTPUT_PATH)

    check_wb = load_workbook(OUTPUT_PATH, data_only=False)
    required_sheets = {"Revision manual notificaciones", "Catalogos", "Instrucciones"}
    missing = required_sheets.difference(check_wb.sheetnames)
    if missing:
        raise RuntimeError(f"Faltan hojas: {sorted(missing)}")
    headers = [
        cell.value
        for cell in check_wb["Revision manual notificaciones"][2]
        if cell.value is not None
    ]
    if headers != HEADERS:
        raise RuntimeError("Los encabezados no coinciden con el esquema esperado")
    print(str(OUTPUT_PATH.resolve()))


if __name__ == "__main__":
    main()
