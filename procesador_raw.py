from io import BytesIO
from typing import Any

from procesador import (
    DETAIL_HEADER_REQUIRED_TERMS,
    NOTIFICATION_METADATA_FIELDS,
    ROW_METADATA_FIELDS,
    add_grouped_date_columns,
    add_grouped_email_columns,
    build_base_case_columns,
    build_notification_columns,
    build_standard_column_dictionary,
    compact_locations,
    extract_date_from_text,
    extract_room_name,
    first_coordinate,
    get_column_index,
    label_matches_required_term,
    normalize_column_name,
    normalize_db_string,
    normalize_label,
    normalize_sheet_name,
    normalize_wide_header,
    pick_metadata,
    serialize_cell_value,
    serialize_date_like_value,
    serialize_text_value,
    should_skip_notification_email,
    split_notification_emails,
    validate_sheet_structure,
)


RAW_ORIGIN = "RAW_INPUT_SALAS"
RAW_HEADER_SCAN_ROWS = 15
RAW_HEADER_SCAN_COLS = 80
RAW_BLANK_ROW_STOP = 100


def payload_allows_raw_processing(payload: dict) -> bool:
    raw_sources = " ".join(
        str(payload.get(field) or "")
        for field in (
            "ruta_sharepoint",
            "carpeta_origen",
            "identifier",
            "nombre_archivo",
        )
    )
    normalized_source = normalize_db_string(raw_sources) or ""
    return "raw" in normalized_source and "input_salas" in normalized_source


def excel_column_coordinate(column_index: int, row_number: int) -> str:
    from openpyxl.utils.cell import get_column_letter

    return f"{get_column_letter(column_index)}{row_number}"


def raw_sheet_value(sheet: dict, row_number: int, column_index: int) -> object:
    rows = sheet["rows"]
    if row_number < 1 or row_number > len(rows):
        return None

    row = rows[row_number - 1]
    if column_index < 1 or column_index > len(row):
        return None

    return row[column_index - 1]


def raw_row_score(values: list[object]) -> int:
    labels = [normalize_label(value) for value in values]
    non_empty = [label for label in labels if label]
    if not non_empty:
        return 0

    required_hits = sum(
        1
        for term in DETAIL_HEADER_REQUIRED_TERMS
        if any(label_matches_required_term(label, term) for label in labels)
    )
    notification_hits = sum(
        1
        for label in labels
        if "CORREO" in label or "GUIA" in label or "FECHA" in label
    )
    return required_hits * 12 + notification_hits * 3 + min(len(non_empty), 30)


def raw_sheet_to_dict_from_openpyxl(worksheet) -> dict:
    max_row = worksheet.max_row or 0
    max_column = min(worksheet.max_column or 0, RAW_HEADER_SCAN_COLS)
    rows = [
        [
            worksheet.cell(row=row_number, column=column_index).value
            for column_index in range(1, max_column + 1)
        ]
        for row_number in range(1, max_row + 1)
    ]
    merges = []
    for merged_range in worksheet.merged_cells.ranges:
        value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        merges.append(
            {
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
                "min_col": merged_range.min_col,
                "max_col": min(merged_range.max_col, max_column),
                "value": value,
            }
        )

    return {
        "name": worksheet.title,
        "rows": rows,
        "max_row": max_row,
        "max_column": max_column,
        "merges": merges,
    }


def load_raw_xls_sheets(content: bytes) -> list[dict]:
    import xlrd

    workbook = xlrd.open_workbook(
        file_contents=content,
        formatting_info=True,
        on_demand=True,
    )
    sheets = []
    for worksheet in workbook.sheets():
        max_column = min(worksheet.ncols, RAW_HEADER_SCAN_COLS)
        rows = [
            worksheet.row_values(row_index, 0, max_column)
            for row_index in range(worksheet.nrows)
        ]
        merges = [
            {
                "min_row": row_start + 1,
                "max_row": row_end,
                "min_col": col_start + 1,
                "max_col": min(col_end, max_column),
                "value": worksheet.cell_value(row_start, col_start),
            }
            for row_start, row_end, col_start, col_end in worksheet.merged_cells
            if col_start < max_column
        ]
        sheets.append(
            {
                "name": worksheet.name,
                "rows": rows,
                "max_row": worksheet.nrows,
                "max_column": max_column,
                "merges": merges,
            }
        )

    return sheets


def find_raw_detail_header_row(sheet: dict) -> int | None:
    max_scan_row = min(sheet["max_row"], RAW_HEADER_SCAN_ROWS)
    best_row = None
    best_score = 0
    for row_number in range(1, max_scan_row + 1):
        score = raw_row_score(sheet["rows"][row_number - 1])
        if score > best_score:
            best_score = score
            best_row = row_number

    return best_row if best_score >= 45 else None


def raw_wide_header_locations(sheet: dict, detail_row: int) -> dict[str, str]:
    locations = {}
    min_row = max(1, detail_row - 6)
    max_row = detail_row - 1

    for merged_range in sheet["merges"]:
        if not (min_row <= merged_range["min_row"] <= max_row):
            continue
        group = normalize_wide_header(merged_range["value"])
        if group:
            locations.setdefault(
                group,
                excel_column_coordinate(merged_range["min_col"], merged_range["min_row"]),
            )

    for row_number in range(min_row, max_row + 1):
        for column_index in range(1, sheet["max_column"] + 1):
            group = normalize_wide_header(raw_sheet_value(sheet, row_number, column_index))
            if group:
                locations.setdefault(group, excel_column_coordinate(column_index, row_number))

    return dict(sorted(locations.items(), key=lambda item: get_column_index(item[1])))


def raw_detail_header_locations(sheet: dict, detail_row: int) -> dict:
    locations = {}
    for column_index, value in enumerate(sheet["rows"][detail_row - 1], start=1):
        label = normalize_label(value)
        if not label:
            for row_number in range(detail_row - 1, max(0, detail_row - 6), -1):
                inherited_label = normalize_label(
                    raw_sheet_value(sheet, row_number, column_index)
                )
                if inherited_label:
                    label = inherited_label
                    break
        if not label:
            continue
        key = normalize_column_name(label)
        locations.setdefault(key, []).append(excel_column_coordinate(column_index, detail_row))

    return compact_locations(locations)


def raw_sheet_metadata(sheet: dict, file_name: str) -> dict | None:
    detail_row = find_raw_detail_header_row(sheet)
    if not detail_row:
        return None

    date_result = extract_date_from_text(sheet["name"]) or extract_date_from_text(file_name)
    if date_result:
        extracted_date, date_span, raw_date = date_result
        room_name = extract_room_name(sheet["name"], date_span)
    else:
        extracted_date = None
        raw_date = None
        room_name = normalize_sheet_name(sheet["name"])

    wide_locations = raw_wide_header_locations(sheet, detail_row)
    detail_locations = raw_detail_header_locations(sheet, detail_row)
    detail_locations = add_grouped_date_columns(detail_locations, wide_locations)
    detail_locations = add_grouped_email_columns(detail_locations, wide_locations)

    sheet_metadata = {
        "pestana_nombre": sheet["name"],
        "pestana_nombre_normalizado": normalize_db_string(sheet["name"]),
        "pestana_fecha": extracted_date,
        "pestana_fecha_original": raw_date,
        "pestana_sala_original": room_name,
        "pestana_sala_normalizada": normalize_db_string(room_name),
        "hoja_trabajo_sala_original": None,
        "hoja_trabajo_sala": None,
        "hoja_trabajo_sala_normalizada": None,
        "hoja_trabajo_sala_celda": None,
        "hoja_trabajo_fecha_audiencia_original": raw_date,
        "hoja_trabajo_fecha_audiencia": extracted_date,
        "hoja_trabajo_fecha_audiencia_celda": None,
        "fila_encabezado_wide": None,
        "fila_encabezado_wide_terminos": sorted(wide_locations),
        "ubicacion_columnas_wide": wide_locations,
        "fila_encabezado_detalle": detail_row,
        "fila_encabezado_detalle_terminos": sorted(detail_locations),
        "ubicacion_columnas_detalle": detail_locations,
        "modo_procesamiento": RAW_ORIGIN,
    }
    sheet_metadata.update(validate_sheet_structure(sheet_metadata, threshold=0.40))
    sheet_metadata["estructura_valida"] = bool(
        detail_locations.get("numero_radicado")
        and build_notification_columns(detail_locations)
    )
    return sheet_metadata


def raw_iter_values_by_coordinate(sheet: dict, start_row: int, coordinates: list[str]):
    column_indexes = {
        coordinate: get_column_index(coordinate)
        for coordinate in coordinates
    }
    blank_streak = 0
    for row_number in range(start_row, sheet["max_row"] + 1):
        values = {
            coordinate: raw_sheet_value(sheet, row_number, column_index)
            for coordinate, column_index in column_indexes.items()
        }
        if all(value in (None, "") for value in values.values()):
            blank_streak += 1
            if blank_streak >= RAW_BLANK_ROW_STOP:
                break
            continue

        blank_streak = 0
        yield row_number, values


def extract_raw_base_case_table(
    raw_sheets: list[dict],
    sheets_metadata: list[dict],
) -> list[dict]:
    rows = []
    sheet_by_name = {sheet["name"]: sheet for sheet in raw_sheets}

    for sheet_metadata in sheets_metadata:
        if not sheet_metadata.get("estructura_valida"):
            continue

        sheet = sheet_by_name[sheet_metadata["pestana_nombre"]]
        detail_row = sheet_metadata["fila_encabezado_detalle"]
        detail_locations = sheet_metadata["ubicacion_columnas_detalle"]
        base_columns = build_base_case_columns(detail_locations)
        key_coordinate = base_columns.get("numero_radicado")
        if not key_coordinate:
            continue

        coordinates = list(base_columns.values())
        for row_number, values_by_coordinate in raw_iter_values_by_coordinate(
            sheet,
            detail_row + 1,
            coordinates,
        ):
            numero_radicado = values_by_coordinate.get(key_coordinate)
            if numero_radicado in (None, ""):
                continue

            output_row = pick_metadata(sheet_metadata, ROW_METADATA_FIELDS)
            output_row.update(
                {
                    "pestana_fecha": sheet_metadata.get("pestana_fecha"),
                    "hoja_trabajo_fecha_audiencia": sheet_metadata.get(
                        "hoja_trabajo_fecha_audiencia"
                    ),
                    "origen_tabla": RAW_ORIGIN,
                    "numero_fila_excel": row_number,
                }
            )
            for column_name, coordinate in base_columns.items():
                output_row[column_name] = serialize_cell_value(
                    values_by_coordinate.get(coordinate)
                )
            rows.append(output_row)

    return rows


def extract_raw_notification_table(
    raw_sheets: list[dict],
    sheets_metadata: list[dict],
) -> list[dict]:
    rows = []
    sheet_by_name = {sheet["name"]: sheet for sheet in raw_sheets}

    for sheet_metadata in sheets_metadata:
        if not sheet_metadata.get("estructura_valida"):
            continue

        sheet = sheet_by_name[sheet_metadata["pestana_nombre"]]
        detail_row = sheet_metadata["fila_encabezado_detalle"]
        detail_locations = sheet_metadata["ubicacion_columnas_detalle"]
        key_coordinate = detail_locations.get("numero_radicado")
        cedula_coordinate = detail_locations.get("cedula")
        notification_columns = build_notification_columns(detail_locations)
        if not key_coordinate or not notification_columns:
            continue

        all_coordinates = [first_coordinate(key_coordinate)]
        if cedula_coordinate:
            all_coordinates.append(first_coordinate(cedula_coordinate))
        for entity_columns in notification_columns.values():
            all_coordinates.extend(entity_columns.values())

        for row_number, values_by_coordinate in raw_iter_values_by_coordinate(
            sheet,
            detail_row + 1,
            all_coordinates,
        ):
            numero_radicado = values_by_coordinate.get(first_coordinate(key_coordinate))
            if numero_radicado in (None, ""):
                continue

            for entity, entity_columns in notification_columns.items():
                output_row = {
                    **pick_metadata(sheet_metadata, NOTIFICATION_METADATA_FIELDS),
                    "numero_radicado": serialize_text_value(numero_radicado),
                    "cedula": serialize_text_value(
                        values_by_coordinate.get(first_coordinate(cedula_coordinate))
                    )
                    if cedula_coordinate
                    else None,
                    "entidad": entity,
                    "fecha_envio": None,
                    "fecha_recibido": None,
                    "correo": None,
                    "origen_tabla": RAW_ORIGIN,
                    "numero_fila_excel": row_number,
                }

                for field, coordinate in entity_columns.items():
                    value = values_by_coordinate.get(coordinate)
                    if field in {"fecha_envio", "fecha_recibido"}:
                        output_row[field] = serialize_date_like_value(value)
                    elif field == "correo":
                        output_row[field] = serialize_text_value(value)
                    else:
                        output_row[field] = serialize_cell_value(value)

                for correo in split_notification_emails(output_row["correo"]):
                    if should_skip_notification_email(correo):
                        continue

                    notification_row = dict(output_row)
                    notification_row["correo"] = correo
                    rows.append(notification_row)

    return rows


def process_raw_excel_data(content: bytes, file_name: str, file_kind: str) -> dict:
    if file_kind == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
        raw_sheets = [
            raw_sheet_to_dict_from_openpyxl(worksheet)
            for worksheet in workbook.worksheets
        ]
    else:
        raw_sheets = load_raw_xls_sheets(content)

    sheets_metadata = [
        metadata
        for sheet in raw_sheets
        if (metadata := raw_sheet_metadata(sheet, file_name)) is not None
    ]
    valid_sheets = [sheet for sheet in sheets_metadata if sheet.get("estructura_valida")]
    base_case_table = extract_raw_base_case_table(raw_sheets, valid_sheets)
    notification_table = extract_raw_notification_table(raw_sheets, valid_sheets)

    return {
        "modo_procesamiento": RAW_ORIGIN,
        "tabla_casos": base_case_table,
        "tabla_notificaciones": notification_table,
        "hojas_con_fecha": sheets_metadata,
        "mensaje_error": [],
        "diccionario_estandar_columnas": build_standard_column_dictionary(sheets_metadata),
    }


def try_process_raw_payload(
    payload: dict[str, Any],
    content: bytes,
    file_kind: str,
    status: str,
    base_case_table: list[dict],
    notification_table: list[dict],
) -> dict | None:
    should_try_raw = (
        status != "OK"
        or (not base_case_table and not notification_table)
    )
    if not should_try_raw:
        return None

    raw_result = process_raw_excel_data(content, payload["nombre_archivo"], file_kind)
    if raw_result["tabla_casos"] or raw_result["tabla_notificaciones"]:
        return raw_result

    return None
